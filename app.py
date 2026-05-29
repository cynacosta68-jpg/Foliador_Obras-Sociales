import streamlit as st
import fitz
import pytesseract
from PIL import Image, ImageEnhance, ImageFilter
import numpy as np
import re
import io
import math
import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
import tempfile, os, base64, time
import urllib.request, urllib.error

st.set_page_config(page_title="Foliador de Obras Sociales", page_icon="📋", layout="centered", initial_sidebar_state="collapsed")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:ital,opsz,wght@0,9..40,300;0,9..40,400;0,9..40,500;0,9..40,600;1,9..40,300&display=swap');
html,body,[class*="css"]{font-family:'DM Sans',sans-serif!important}
.stApp{background:linear-gradient(160deg,#060e1a 0%,#0a1628 40%,#0f1f3d 100%)}
.main-header{text-align:center;padding:2.5rem 0 1rem}
.main-header h1{font-family:'DM Sans',sans-serif!important;font-weight:600;font-size:2rem;color:#e2e8f0;letter-spacing:-.5px;margin-bottom:.25rem}
.main-header p{font-family:'DM Sans',sans-serif!important;font-weight:300;font-size:.95rem;color:#64748b;letter-spacing:.3px}
.accent-line{width:48px;height:2px;background:#3b82f6;margin:.75rem auto;border-radius:2px}
.card{background:rgba(17,29,51,.6);border:1px solid rgba(59,130,246,.08);border-radius:12px;padding:1.75rem;margin-bottom:1.25rem;backdrop-filter:blur(8px)}
.card-title{font-family:'DM Sans',sans-serif!important;font-weight:500;font-size:.8rem;color:#60a5fa;text-transform:uppercase;letter-spacing:1.5px;margin-bottom:1rem}
[data-testid="stFileUploader"]{background:rgba(15,23,42,.5);border:1px dashed rgba(59,130,246,.2);border-radius:10px;padding:1rem}
.stButton>button{font-family:'DM Sans',sans-serif!important;font-weight:500;background:linear-gradient(135deg,#2563eb,#3b82f6)!important;color:#fff!important;border:none!important;border-radius:8px!important;padding:.6rem 2rem!important;transition:all .2s}
.stButton>button:hover{background:linear-gradient(135deg,#1d4ed8,#2563eb)!important;box-shadow:0 4px 20px rgba(59,130,246,.3)!important}
.stDownloadButton>button{font-family:'DM Sans',sans-serif!important;font-weight:500;background:linear-gradient(135deg,#059669,#10b981)!important;color:#fff!important;border:none!important;border-radius:8px!important;padding:.6rem 2rem!important;width:100%}
.stat-row{display:flex;gap:1rem;margin:1rem 0;flex-wrap:wrap}
.stat-box{flex:1;min-width:100px;background:rgba(15,23,42,.5);border:1px solid rgba(59,130,246,.1);border-radius:10px;padding:1.1rem;text-align:center}
.stat-number{font-family:'DM Sans',sans-serif!important;font-size:1.8rem;font-weight:600;color:#60a5fa;line-height:1}
.stat-label{font-family:'DM Sans',sans-serif!important;font-size:.7rem;color:#64748b;text-transform:uppercase;letter-spacing:1px;margin-top:.4rem}
.code-card{background:rgba(15,23,42,.5);border:1px solid rgba(59,130,246,.12);border-radius:8px;padding:.6rem;text-align:center;margin-bottom:.5rem}
.code-card img{border-radius:4px;width:100%;background:#fff;padding:4px}
.page-num{font-family:'DM Sans',sans-serif;font-size:.7rem;color:#94a3b8;margin-top:.3rem}
.ocr-tag{display:inline-block;background:rgba(59,130,246,.15);color:#60a5fa;font-family:'DM Sans',sans-serif;font-size:.75rem;font-weight:500;padding:.15rem .5rem;border-radius:4px;margin-top:.25rem}
.no-tag{display:inline-block;background:rgba(251,191,36,.1);color:#fbbf24;font-family:'DM Sans',sans-serif;font-size:.7rem;padding:.15rem .5rem;border-radius:4px;margin-top:.25rem}
.match-ok{display:inline-block;background:rgba(16,185,129,.12);color:#34d399;font-family:'DM Sans',sans-serif;font-size:.65rem;font-weight:500;padding:.12rem .4rem;border-radius:3px;margin-top:.2rem}
.match-fail{display:inline-block;background:rgba(239,68,68,.12);color:#f87171;font-family:'DM Sans',sans-serif;font-size:.65rem;font-weight:500;padding:.12rem .4rem;border-radius:3px;margin-top:.2rem}
.engine-tag{display:inline-block;background:rgba(168,85,247,.12);color:#c084fc;font-family:'DM Sans',sans-serif;font-size:.6rem;padding:.1rem .4rem;border-radius:3px;margin-left:.3rem}
.stAlert{border-radius:10px!important}
#MainMenu{visibility:hidden}footer{visibility:hidden}
.info-text{font-family:'DM Sans',sans-serif;font-size:.82rem;color:#94a3b8;margin-bottom:1rem;line-height:1.5}
.step-badge{display:inline-block;background:rgba(59,130,246,.12);color:#60a5fa;font-family:'DM Sans',sans-serif;font-size:.65rem;font-weight:600;padding:.2rem .6rem;border-radius:20px;letter-spacing:1px;text-transform:uppercase;margin-bottom:.5rem}
</style>
""", unsafe_allow_html=True)

CODE_RE = re.compile(r'[CcĆć]\s*(\d{1,5})')

# ═══════════════════════════════════════════════════════════════
# AUTENTICACIÓN
# ═══════════════════════════════════════════════════════════════

def check_password():
    """Gate de acceso con contraseña. Retorna True si autenticado."""
    try:
        correct_pw = st.secrets["PASSWORD"]
    except Exception:
        return True  # Sin PASSWORD en secrets → acceso libre

    if st.session_state.get("authenticated"):
        return True

    st.markdown("""
    <div class="main-header">
        <h1>📋 Foliador de Obras Sociales</h1>
        <div class="accent-line"></div>
        <p>Ingresá la contraseña para acceder</p>
    </div>
    """, unsafe_allow_html=True)

    st.markdown('<div class="card" style="max-width:380px;margin:0 auto;">', unsafe_allow_html=True)
    pw = st.text_input("Contraseña", type="password", placeholder="Ingresá la contraseña...")
    if st.button("Ingresar", use_container_width=True):
        if pw == correct_pw:
            st.session_state["authenticated"] = True
            st.rerun()
        else:
            st.error("Contraseña incorrecta.")
    st.markdown('</div>', unsafe_allow_html=True)
    return False

if not check_password():
    st.stop()

# ═══════════════════════════════════════════════════════════════
# API KEY
# ═══════════════════════════════════════════════════════════════

def get_api_key():
    try:
        return st.secrets["ANTHROPIC_API_KEY"]
    except Exception:
        return os.environ.get("ANTHROPIC_API_KEY", "")

# ═══════════════════════════════════════════════════════════════
# BASE DE USUARIOS
# ═══════════════════════════════════════════════════════════════

@st.cache_data
def load_base(file_bytes):
    df = pd.read_excel(io.BytesIO(file_bytes))
    df.columns = df.columns.str.strip()
    lookup = {}
    for _, row in df.iterrows():
        mat = str(row.get('Matricula', '')).strip()
        if mat:
            lookup[mat] = {
                'Nombre': str(row.get('Nombre', '')).strip(),
                'CUIT': row.get('CUIT', ''),
                'Especialidad': str(row.get('Especialidad', '')).strip(),
                'Responsabilidad Fiscal': str(row.get('Responsabilidad Fiscal', '')).strip(),
                'Arancel': str(row.get('Arancel', '')).strip(),
            }
    return lookup, len(df)

def do_lookup(code, base):
    if not base or not code: return None
    code = code.strip()
    return base.get(code)

# ═══════════════════════════════════════════════════════════════
# OCR ENGINES
# ═══════════════════════════════════════════════════════════════

def claude_vision_read(img_b64, api_key):
    """Use Claude Vision to read a handwritten code from an image."""
    body = json.dumps({
        "model": "claude-haiku-4-5-20251001",
        "max_tokens": 30,
        "messages": [{
            "role": "user",
            "content": [
                {"type": "image", "source": {"type": "base64", "media_type": "image/png", "data": img_b64}},
                {"type": "text", "text": "Lee el código manuscrito de esta imagen. Es formato C seguido de números (ej: C1098, C50, C294). Respondé SOLO el código, nada más."}
            ]
        }]
    })
    req = urllib.request.Request(
        "https://api.anthropic.com/v1/messages",
        data=body.encode(),
        headers={"Content-Type": "application/json", "x-api-key": api_key, "anthropic-version": "2023-06-01"}
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read())
            text = data.get("content", [{}])[0].get("text", "").strip()
            m = re.match(r'(C\d{1,5})', text)
            return m.group(1) if m else text if text.startswith('C') else None
    except Exception:
        return None


def tesseract_read(gray_img):
    """Tesseract multi-strategy OCR (v1 approach)."""
    results = []
    for thresh in [90, 110, 130, 150, 170]:
        binary = gray_img.point(lambda x, t=thresh: 0 if x < t else 255, '1')
        scaled = binary.resize((binary.width * 3, binary.height * 3), Image.LANCZOS)
        for psm in [6, 7, 11, 13]:
            for cfg in [f'--psm {psm}', f'--psm {psm} -c tessedit_char_whitelist=Cc0123456789']:
                try:
                    text = pytesseract.image_to_string(scaled, config=cfg)
                    for m in CODE_RE.findall(text):
                        results.append(f"C{m}")
                except Exception:
                    pass
    eroded = gray_img.filter(ImageFilter.MaxFilter(3))
    be = eroded.point(lambda x: 0 if x < 130 else 255, '1')
    se = be.resize((be.width * 3, be.height * 3), Image.LANCZOS)
    for psm in [6, 7]:
        try:
            text = pytesseract.image_to_string(se, config=f'--psm {psm}')
            for m in CODE_RE.findall(text):
                results.append(f"C{m}")
        except Exception:
            pass
    if results:
        best, cnt = Counter(results).most_common(1)[0]
        if cnt >= 2:
            return best
    return None


# ═══════════════════════════════════════════════════════════════
# SCAN ENGINE
# ═══════════════════════════════════════════════════════════════

def phase1_density(pdf_path, total, progress_cb=None):
    doc = fitz.open(pdf_path)
    densities = []
    for i in range(total):
        page = doc[i]
        pix = page.get_pixmap(matrix=fitz.Matrix(1, 1))
        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        crop = img.crop((0, 0, int(img.width * 0.25), int(img.height * 0.06)))
        densities.append(float(np.mean(np.array(crop.convert('L')) < 100)))
        if progress_cb and i % 100 == 0:
            progress_cb(0.25 * i / total, f"Fase 1 · Densidad · Pág {i+1:,}/{total:,}")
    doc.close()
    median = float(np.median(densities)) if densities else 0
    return [i for i, d in enumerate(densities) if d > max(median * 3, 0.02)]


def _process_candidate(pdf_path, pg, api_key):
    """Process one candidate: text extraction → Claude Vision → Tesseract fallback."""
    doc = fitz.open(pdf_path)
    page = doc[pg]

    # 1. Text extraction (instant)
    for blk in page.get_text("blocks"):
        if blk[1] < page.rect.height * 0.15 and blk[0] < page.rect.width * 0.40:
            m = CODE_RE.findall(blk[4] if len(blk) > 4 else "")
            if m:
                doc.close()
                return pg, f"C{m[0]}", "text"

    # 2. Render crop
    mat = fitz.Matrix(250 / 72, 250 / 72)
    pix = page.get_pixmap(matrix=mat)
    img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
    doc.close()
    crop = img.crop((0, 0, int(img.width * 0.35), int(img.height * 0.18)))

    # 3. Claude Vision (if API key available)
    if api_key:
        buf = io.BytesIO()
        crop.save(buf, format='PNG')
        b64 = base64.b64encode(buf.getvalue()).decode()
        code = claude_vision_read(b64, api_key)
        if code:
            return pg, code, "claude"

    # 4. Tesseract fallback
    gray = crop.convert('L')
    code = tesseract_read(gray)
    if code:
        return pg, code, "tesseract"

    return pg, None, None


def full_scan(pdf_path, total, api_key="", progress_cb=None):
    if progress_cb: progress_cb(0.0, "Fase 1 · Detección por densidad...")
    candidates = phase1_density(pdf_path, total, progress_cb)
    if progress_cb: progress_cb(0.3, f"Fase 1 · {len(candidates)} candidatas")

    results = {}
    if not candidates:
        return results, candidates

    engine = "Claude Vision" if api_key else "Tesseract OCR"
    workers = min(4, os.cpu_count() or 2) if not api_key else 2  # Less parallel for API
    done = 0
    if progress_cb: progress_cb(0.35, f"Fase 2 · {engine} en {len(candidates)} candidatas...")

    with ThreadPoolExecutor(max_workers=workers) as ex:
        futures = {ex.submit(_process_candidate, pdf_path, pg, api_key): pg for pg in candidates}
        for f in as_completed(futures):
            pg, code, method = f.result()
            if code: results[pg] = (code, method)
            done += 1
            if progress_cb and done % 2 == 0:
                progress_cb(0.35 + 0.6 * done / len(candidates), f"Fase 2 · {done}/{len(candidates)}")

    if progress_cb: progress_cb(1.0, "Completado")
    return results, candidates


def get_thumb_b64(doc, pg, width=220):
    page = doc[pg]
    pix = page.get_pixmap(matrix=fitz.Matrix(150 / 72, 150 / 72))
    img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
    crop = img.crop((0, 0, img.width, int(img.height * 0.20)))
    r = width / crop.width
    crop = crop.resize((width, int(crop.height * r)), Image.LANCZOS)
    buf = io.BytesIO()
    crop.save(buf, format='JPEG', quality=55)
    return base64.b64encode(buf.getvalue()).decode()


# ═══════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════

def build_ranges(code_pages, total):
    entries = sorted(code_pages.items())
    return [(code, pg+1, entries[i+1][0] if i+1 < len(entries) else total)
            for i, (pg, (code, _)) in enumerate(entries)]

def apply_lookup(df, base):
    out = df.copy()
    for idx, row in out.iterrows():
        code = str(row.get('Cuenta', '')).strip()
        info = do_lookup(code, base)
        if info:
            out.at[idx, 'Profesional'] = info['Nombre']
            out.at[idx, 'CUIT'] = str(int(info['CUIT'])) if pd.notna(info.get('CUIT')) else ""
            out.at[idx, 'Especialidad'] = info['Especialidad']
            out.at[idx, 'Resp. Fiscal'] = info['Responsabilidad Fiscal']
            out.at[idx, 'Arancel'] = info['Arancel']
        else:
            out.at[idx, 'Profesional'] = ""
            out.at[idx, 'CUIT'] = ""
            out.at[idx, 'Especialidad'] = ""
            out.at[idx, 'Resp. Fiscal'] = ""
            out.at[idx, 'Arancel'] = ""
    return out

def create_excel(rows, headers, widths):
    wb = Workbook(); ws = wb.active; ws.title = "Foliado"
    hf = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    hfill = PatternFill('solid', fgColor='1a3a6b')
    ha = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cf = Font(name='Calibri', size=11)
    ac, al = Alignment(horizontal='center', vertical='center'), Alignment(horizontal='left', vertical='center')
    bdr = Border(left=Side(style='thin',color='2c5282'), right=Side(style='thin',color='2c5282'),
                 top=Side(style='thin',color='2c5282'), bottom=Side(style='thin',color='2c5282'))
    alt = PatternFill('solid', fgColor='edf2f7')
    left_cols = {'Profesional', 'Especialidad', 'Resp. Fiscal'}
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font, c.fill, c.alignment, c.border = hf, hfill, ha, bdr
        ws.column_dimensions[get_column_letter(ci)].width = w
    for ri, rd in enumerate(rows, 2):
        vals = (list(rd) + [''] * len(headers))[:len(headers)]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font, c.border = cf, bdr
            c.alignment = al if headers[ci-1] in left_cols else ac
            if ri % 2 == 0: c.fill = alt
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows)+1}"
    ws.freeze_panes = "A2"
    buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf

# ═══════════════════════════════════════════════════════════════
# UI
# ═══════════════════════════════════════════════════════════════

st.markdown("""
<div class="main-header">
    <h1>📋 Foliador de Obras Sociales</h1>
    <div class="accent-line"></div>
    <p>Extrae códigos de cuenta, cruza con la base de profesionales y genera el Excel</p>
</div>
""", unsafe_allow_html=True)

# ── PASO 1 ──
st.markdown('<div class="card"><span class="step-badge">Paso 1</span><div class="card-title">Cargar documentos</div>', unsafe_allow_html=True)
col_pdf, col_base = st.columns(2)
with col_pdf:
    st.markdown('<p class="info-text">📄 <b>PDF escaneado</b></p>', unsafe_allow_html=True)
    uploaded_pdf = st.file_uploader("PDF", type=["pdf"], label_visibility="collapsed", key="pdf")
with col_base:
    st.markdown('<p class="info-text">📊 <b>Base de usuarios</b> (.xlsx)</p>', unsafe_allow_html=True)
    uploaded_base = st.file_uploader("Base", type=["xlsx","xls"], label_visibility="collapsed", key="base")

# API Key
api_key = get_api_key()
if not api_key:
    with st.expander("🔑 API Key de Anthropic (opcional pero recomendado)"):
        st.markdown('<p class="info-text">Con la API Key, la app usa <b>Claude Vision</b> para leer los códigos manuscritos con mucha mayor precisión que Tesseract OCR. Sin ella, funciona igual pero con OCR tradicional.</p>', unsafe_allow_html=True)
        api_key = st.text_input("Anthropic API Key", type="password", placeholder="sk-ant-...")
        st.markdown('<p class="info-text">Para Streamlit Cloud: agregala en Settings → Secrets como <code>ANTHROPIC_API_KEY = "sk-ant-..."</code></p>', unsafe_allow_html=True)

st.markdown('</div>', unsafe_allow_html=True)

base_lookup, base_count = {}, 0
if uploaded_base:
    base_lookup, base_count = load_base(uploaded_base.read()); uploaded_base.seek(0)

if uploaded_pdf:
    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
        tmp.write(uploaded_pdf.read()); tmp_path = tmp.name
    doc = fitz.open(tmp_path); total = len(doc); doc.close()

    engine = "Claude Vision 🧠" if api_key else "Tesseract OCR"
    stats = f'<div class="stat-row"><div class="stat-box"><div class="stat-number">{total:,}</div><div class="stat-label">Páginas</div></div><div class="stat-box"><div class="stat-number">{uploaded_pdf.size/1024/1024:.1f} MB</div><div class="stat-label">Tamaño</div></div>'
    if base_lookup: stats += f'<div class="stat-box"><div class="stat-number">{base_count:,}</div><div class="stat-label">En base</div></div>'
    stats += f'<div class="stat-box"><div class="stat-number">{engine}</div><div class="stat-label">Motor</div></div></div>'
    st.markdown(stats, unsafe_allow_html=True)

    # ── PASO 2 ──
    st.markdown('<div class="card"><span class="step-badge">Paso 2</span><div class="card-title">Escaneo + cruce</div>', unsafe_allow_html=True)

    if st.button("🔍  Escanear y cruzar", use_container_width=True):
        bar = st.progress(0)
        t0 = time.time()
        results, candidates = full_scan(tmp_path, total, api_key, lambda p, m: bar.progress(min(p,1.0), text=m))
        elapsed = time.time() - t0
        bar.empty()
        st.session_state.update({'results': results, 'candidates': candidates, 'scanned': True, 'pdf_path': tmp_path})

        matches = sum(1 for _, (c, _) in results.items() if do_lookup(c, base_lookup))
        st.markdown(f'<div class="stat-row"><div class="stat-box"><div class="stat-number">{len(candidates)}</div><div class="stat-label">Candidatas</div></div><div class="stat-box"><div class="stat-number">{len(results)}</div><div class="stat-label">Leídos</div></div><div class="stat-box"><div class="stat-number">{matches}/{len(results)}</div><div class="stat-label">Cruzados</div></div><div class="stat-box"><div class="stat-number">{elapsed:.1f}s</div><div class="stat-label">Tiempo</div></div></div>', unsafe_allow_html=True)
        no_match = len(results) - matches
        if matches > 0 and no_match == 0:
            st.success(f"**{matches}** códigos detectados y **todos cruzados** con la base.")
        elif matches > 0:
            st.success(f"**{matches}** cruzados. **{no_match}** sin match — corregí en la tabla y re-cruzá.")
        elif results:
            st.warning("Códigos detectados pero sin match. Corregí en la tabla y re-cruzá.")
        else:
            st.warning("No se detectaron códigos. Ingresalos manualmente.")

    if st.session_state.get('scanned'):
        results = st.session_state['results']
        candidates = st.session_state['candidates']
        if candidates:
            st.markdown("**Páginas con código detectado:**")
            doc2 = fitz.open(st.session_state.get('pdf_path', tmp_path))
            per_page = 12
            total_groups = math.ceil(len(candidates)/per_page)
            group = st.number_input("Grupo",1,total_groups,1) if total_groups > 1 else 1
            vis = candidates[(group-1)*per_page:group*per_page]
            for rs in range(0, len(vis), 4):
                cols = st.columns(4)
                for j, col in enumerate(cols):
                    if rs+j >= len(vis): break
                    pg = vis[rs+j]
                    with col:
                        try:
                            thumb = get_thumb_b64(doc2, pg)
                            if pg in results:
                                code, method = results[pg]
                                info = do_lookup(code, base_lookup)
                                etag = f'<span class="engine-tag">{method}</span>'
                                tag = f'<span class="ocr-tag">{code}</span>{etag}'
                                if info:
                                    n = info["Nombre"]; n = n[:22]+"…" if len(n)>22 else n
                                    mtag = f'<br><span class="match-ok">✓ {n}</span>'
                                else:
                                    mtag = '<br><span class="match-fail">✗ sin match</span>'
                            else:
                                tag = '<span class="no-tag">no leído</span>'; mtag = ''
                            st.markdown(f'<div class="code-card"><img src="data:image/jpeg;base64,{thumb}"/><div class="page-num">Pág. {pg+1}</div>{tag}{mtag}</div>', unsafe_allow_html=True)
                        except: st.text(f"Pág. {pg+1}")
            doc2.close()
    st.markdown('</div>', unsafe_allow_html=True)

    # ── PASO 3 ──
    if st.session_state.get('scanned'):
        results = st.session_state['results']
        has_base = bool(base_lookup)
        st.markdown('<div class="card"><span class="step-badge">Paso 3</span><div class="card-title">Confirmar, editar y cruzar</div>', unsafe_allow_html=True)
        st.markdown('<p class="info-text">Corregí los códigos mirando las miniaturas. Después presioná <b>Re-cruzar</b> para actualizar los datos del profesional.</p>', unsafe_allow_html=True)

        if 'table_df' not in st.session_state or st.session_state.get('rebuild_table'):
            if results:
                ranges = build_ranges(results, total)
                rows = []
                for code, s, e in ranges:
                    info = do_lookup(code, base_lookup)
                    rows.append({"Cuenta": code, "Página desde": s, "Página hasta": e,
                                 "Profesional": info['Nombre'] if info else "",
                                 "CUIT": str(int(info['CUIT'])) if info and pd.notna(info.get('CUIT')) else "",
                                 "Especialidad": info['Especialidad'] if info else "",
                                 "Resp. Fiscal": info['Responsabilidad Fiscal'] if info else "",
                                 "Arancel": info['Arancel'] if info else ""})
            else:
                rows = [{"Cuenta":"","Página desde":1,"Página hasta":total,"Profesional":"","CUIT":"","Especialidad":"","Resp. Fiscal":"","Arancel":""}]
            st.session_state['table_df'] = pd.DataFrame(rows)
            st.session_state['rebuild_table'] = False

        edited = st.data_editor(st.session_state['table_df'], use_container_width=True, num_rows="dynamic", hide_index=True,
            column_config={
                "Cuenta": st.column_config.TextColumn("Cuenta", width="small"),
                "Página desde": st.column_config.NumberColumn("Pág. desde", width="small", min_value=1, max_value=total),
                "Página hasta": st.column_config.NumberColumn("Pág. hasta", width="small", min_value=1, max_value=total),
                "Profesional": st.column_config.TextColumn("Profesional", width="medium"),
                "CUIT": st.column_config.TextColumn("CUIT", width="small"),
                "Especialidad": st.column_config.TextColumn("Especialidad", width="medium"),
                "Resp. Fiscal": st.column_config.TextColumn("Resp. Fiscal", width="small"),
                "Arancel": st.column_config.TextColumn("Arancel", width="small"),
            })

        if has_base:
            if st.button("🔄  Re-cruzar con base", use_container_width=True):
                updated = apply_lookup(edited, base_lookup)
                st.session_state['table_df'] = updated
                matched = sum(1 for _, r in updated.iterrows() if str(r.get('Profesional','')).strip())
                st.success(f"**{matched}/{len(updated)}** cruzados con la base.")
                st.rerun()

        st.session_state['final_df'] = edited
        st.markdown('</div>', unsafe_allow_html=True)

        # ── PASO 4 ──
        st.markdown('<div class="card"><span class="step-badge">Paso 4</span><div class="card-title">Descargar Excel</div>', unsafe_allow_html=True)
        headers = ['Cuenta','Página desde','Página hasta','Profesional','CUIT','Especialidad','Resp. Fiscal','Arancel']
        widths = [12,14,14,36,18,28,22,10]
        excel = create_excel(st.session_state['final_df'].values.tolist(), headers, widths)
        fname = uploaded_pdf.name.replace(".pdf","").replace(".PDF","")
        st.download_button("⬇  Descargar Excel", excel, f"Foliado_{fname}.xlsx",
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
        st.markdown('</div>', unsafe_allow_html=True)

st.markdown('<div style="text-align:center;padding:3rem 0 1.5rem;border-top:1px solid rgba(59,130,246,.06);margin-top:2rem"><p style="font-size:.72rem;color:#475569;letter-spacing:.5px;font-family:\'DM Sans\',sans-serif">FOLIADOR DE OBRAS SOCIALES · v4.2 · Claude Haiku Vision + cruce con base + acceso protegido</p></div>', unsafe_allow_html=True)
